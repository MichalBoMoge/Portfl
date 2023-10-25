from bs4 import BeautifulSoup
import requests
import re
import pandas as pd
import openpyxl
import os
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import Workbook

html = requests.get("https://www.keil.com/dd/chips/all/arm.htm")




soup = BeautifulSoup(html.text, 'html.parser')
i = 0
links = soup.find_all('li')

max_len = 0

dfs = []
for link in links:
    if link.find('a', attrs={'class':'vn'}) : 
        tytulowy = link.find('a', attrs={'class':'vn'})
        lista_left = []
        pozostale = link.find_all('a', href= lambda href: href and href.startswith(f'/dd/chip/'))
        if len(pozostale) > max_len : max_len = len(pozostale)
        for nazwy in pozostale:
            lista_left.append(nazwy.text)
        dane = pd.DataFrame(data={tytulowy.text : lista_left})
        dfs.append(dane)
fin = pd.concat(dfs, axis=1)
fin.idxmax = max_len

fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

if os.path.exists("output.xlsx") == False:
    fin.to_excel("output.xlsx")
else:
    dlugosc = len(fin.columns)+1
    df = pd.read_excel("output.xlsx",'Sheet1')
    df = df.iloc[:,1:dlugosc]
    for row in fin.iterrows():
        index, row_data = row
        for column, value in row_data.items():
            if not pd.isna(value):
                fin.at[index,column]  = value
                if pd.isna(df.at[index,column]) : 
                    fin.at[index,column] = "XXX" + value
    fin.to_excel("output.xlsx")

    exc_file = openpyxl.load_workbook('output.xlsx')
    worksheet = exc_file.active
    find = "XXX"

    for column in worksheet.columns:
        worksheet.column_dimensions[column[0].column_letter].width = 25


    clear = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.fill = clear
            cell.border = border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
        for cell in row:
            if cell.value and str(cell.value).startswith(find):
                cell.value = str(cell.value)[3:]
                cell.fill = fill 

    exc_file.save('output.xlsx')


