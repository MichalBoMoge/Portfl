import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill


with open("LLVM_SOURCE_MNEMONICS.txt","r") as plik:
    llvm = plik.readlines()

for i in range (0, len(llvm)):
    llvm[i] = llvm[i].strip()
    llvm[i] = llvm[i].lstrip()
    llvm[i] = llvm[i].rstrip()
    parts = llvm[i].split(" ",1)
    llvm[i] = parts[0] + " "+ "".join(parts[1].split())
    

with open("T32_MNEMONICS.log","r") as plik:
    t32 = plik.readlines()

for i in range (0, len(t32)):
    t32[i] = t32[i].strip()
    t32[i] = t32[i].rstrip()
    t32[i] = t32[i].lstrip()
    parts = t32[i].split(" ",1)
    t32[i] = parts[0] + " "+"".join(parts[1].split())
    t32[i] = t32[i].lower()

test_sve = []
with open("test_sve2p1.cmm") as plik:
    lines = plik.readlines()
    for line in lines:
        if "d.s " in line: 
            bf = line.split(" //")[0]
            test_sve.append(bf)

result = []
dane = []
poprawne = 0
niepoprawne = 0
for i in range(0, len(llvm)):
    if t32[i] == llvm[i] : 
        result.append(True)
        poprawne += 1
    else : 
        result.append(False)
        niepoprawne +=1
        print(t32[i],"xxxxxxxxxxx",llvm[i])
    bufor = [llvm[i], t32[i], test_sve[i], result[i]]
    dane.append(bufor)


df = pd.DataFrame(data=dane, columns=["LLVM instruction name", "Disassembled commamnd", "Command","Result"] )
df.index.name = "test index"


df.to_excel("Tests_Results.xlsx", sheet_name="tests", startrow=1, startcol=1, index=True, header=True)
wb = openpyxl.load_workbook("Tests_Results.xlsx")
ws = wb['tests']
ws.column_dimensions["B"].width = 15

for char in "CDE":
    ws.column_dimensions[char].width = 30
    ws.column_dimensions[char].alignment = Alignment(horizontal="center")

thin = Side(border_style="thin", color="000000")

for i in range(3, len(result)+3):
     
    for char in "BCDE":
        cell= char +str(i)
        if i%2 == 1 :    
            ws[cell].fill = PatternFill(fgColor="0064FF", fill_type="solid")
        else : ws[cell].fill = PatternFill(fgColor="0096FF", fill_type="solid")
        ws[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)


    cell = "F"+str(i)
    if ws[cell].value == True : 
        ws[cell].fill = PatternFill(fgColor="008000" ,fill_type="solid")
        ws[cell].value = "correct"
    else: 
        ws[cell].fill = PatternFill(fgColor="FF0000" ,fill_type="solid") 
        ws[cell].value = "incorrect"
    ws[cell].alignment = Alignment(horizontal="center")
    ws[cell].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("I2:J2")
    komorka = ws.cell(row=2,column=9)
    komorka.value = "Tests Count"
    komorka.alignment = Alignment(horizontal="center")
    komorka.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    komorka.font = Font(size=15, bold=True)
    ws["I3"].value = "Correct"
    ws["I3"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["I3"].fill = PatternFill(fgColor="008000" ,fill_type="solid")
    ws["I4"].value = poprawne
    ws["J3"].value = "Incorrect"
    ws["J3"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["J3"].fill = PatternFill(fgColor="FF0000" ,fill_type="solid") 
    ws["J4"].value = niepoprawne

wb.save("Tests_Results.xlsx")